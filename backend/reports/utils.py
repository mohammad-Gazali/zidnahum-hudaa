from django.contrib.auth import get_user_model
from reports.serializers import (
    ReportMemorizeMessageSerializer, 
    ReportsStudentResponseSerializer, 
    ReportsStudentCategoryOrGroupStudentSerializer, 
    ReportsStudentCategoryOrGroupResponseSerializer,
)
from students.models import MemorizeMessage, MessageTypeChoice, Student, StudentCategory, StudentGroup, StudentMasjedChoice
from students.utils import get_num_pages_memo, get_num_pages_test, display_memorize_message_changes
from collections.abc import Iterable
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font
from io import BytesIO

def get_student_report(student: Student, start_date, end_date):
    messages = student.memorizemessage_set.filter(
        sended_at__date__range=[start_date, end_date],
        message_type__in=[MessageTypeChoice.MEMO, MessageTypeChoice.TEST],
    )

    sum_memo, sum_test = _get_sums_student_report(messages)

    messages_data = ReportMemorizeMessageSerializer(messages, many=True)

    report_serializer = ReportsStudentResponseSerializer({ 
        'messages': messages_data.data,
        'sum_memo': sum_memo,
        'sum_test': sum_test,
        'sum_all': sum_memo + sum_test,
    })

    return report_serializer.data


def get_category_or_group_report(category_or_group: StudentCategory | StudentGroup, masjed, start_date, end_date):
    total_memo = 0
    total_test = 0
    students_result = []

    students = category_or_group.student_set.filter(masjed=masjed)

    for student in students:
        messages = student.memorizemessage_set.filter(
            sended_at__date__range=[start_date, end_date],
            message_type__in=[MessageTypeChoice.MEMO, MessageTypeChoice.TEST],
        )

        sum_memo, sum_test = _get_sums_student_report(messages)

        total_memo += sum_memo
        total_test += sum_test

        students_result.append(ReportsStudentCategoryOrGroupStudentSerializer({
            'student_id': student.pk,
            'student_name': student.name,
            'sum_memo': sum_memo,
            'sum_test': sum_test,
            'sum_all': sum_memo + sum_test,
        }).data)


    return ReportsStudentCategoryOrGroupResponseSerializer({
        'students': students_result,
        'total_memo': total_memo,
        'total_test': total_test,
        'total': total_memo + total_test,
    }).data


def _get_sums_student_report(messages: Iterable[MemorizeMessage]):
    sum_memo = 0
    sum_test = 0

    for message in messages:
        if message.message_type == MessageTypeChoice.MEMO:
            sum_memo += get_num_pages_memo(message.changes)

        elif message.message_type == MessageTypeChoice.TEST:
            sum_test += get_num_pages_test(message.changes)

    return sum_memo, sum_test


def excel_student_report(data, student_name: str, start_date: str, end_date: str):
    workbook = Workbook()

    sheet: Worksheet = workbook.active

    sheet.title = 'التقرير'

    alignment = Alignment(horizontal='center', vertical='center')
    alignment_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_font = Font(bold=True)

    cell = sheet.cell(1, 1, 'الطالب')
    cell.alignment = alignment
    cell.font = header_font
    sheet.cell(1, 2, student_name).alignment = alignment

    cell = sheet.cell(2, 1, 'تاريخ البداية')
    cell.alignment = alignment
    cell.font = header_font
    sheet.cell(2, 2, start_date).alignment = alignment

    cell = sheet.cell(3, 1, 'تاريخ النهاية')
    cell.alignment = alignment
    cell.font = header_font
    sheet.cell(3, 2, end_date).alignment = alignment

    cell = sheet.cell(4, 1, 'صفحات التسميع')
    cell.alignment = alignment
    cell.font = header_font
    sheet.cell(4, 2, data['sum_memo']).alignment = alignment

    cell = sheet.cell(5, 1, 'صفحات السبر')
    cell.alignment = alignment
    cell.font = header_font
    sheet.cell(5, 2, data['sum_test']).alignment = alignment

    cell = sheet.cell(6, 1, 'كلي الصفحات')
    cell.alignment = alignment
    cell.font = header_font
    sheet.cell(6, 2, data['sum_all']).alignment = alignment

    details_cell_start = 8

    cell = sheet.cell(details_cell_start, 1, 'الأستاذ')
    cell.alignment = alignment
    cell.font = header_font

    cell = sheet.cell(details_cell_start, 2, 'المحتوى')
    cell.alignment = alignment
    cell.font = header_font

    cell = sheet.cell(details_cell_start, 3, 'نوع الرسالة')
    cell.alignment = alignment
    cell.font = header_font

    cell = sheet.cell(details_cell_start, 4, 'القيمة مضاعفة')
    cell.alignment = alignment
    cell.font = header_font

    master_map = _get_masters_map()

    for row, message in enumerate(data['messages'], details_cell_start + 1):
        master = master_map.get(message['master'], '-')
        changes = display_memorize_message_changes(message['changes'], message['message_type'])
        message_type = 'تسميع' if message['message_type'] == 1 else 'سبر'
        is_doubled = 'نعم' if message['is_doubled'] else 'لا'

        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 60
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 12

        sheet.cell(row, 1, master).alignment = alignment
        sheet.cell(row, 2, changes).alignment = alignment_wrap
        sheet.cell(row, 3, message_type).alignment = alignment
        sheet.cell(row, 4, is_doubled).alignment = alignment
    
    buffer = BytesIO()

    workbook.save(buffer)

    buffer.seek(0)

    return buffer.getvalue()


def excel_category_or_group_report(data, category_or_group_name: str, masjed: int, start_date: str, end_date: str, is_category: bool):
    workbook = Workbook()

    sheet: Worksheet = workbook.active

    sheet.title = 'التقرير'

    alignment = Alignment(horizontal='center', vertical='center') 
    header_font = Font(bold=True)

    cell = sheet.cell(1, 1, 'الفئة' if is_category else 'المجموعة')
    cell.alignment = alignment
    cell.font = header_font
    sheet.cell(1, 2, category_or_group_name).alignment = alignment

    cell = sheet.cell(2, 1, 'المسجد')
    cell.alignment = alignment
    cell.font = header_font
    if masjed == StudentMasjedChoice.HASANIN:
        sheet.cell(2, 2, 'الحسنين').alignment = alignment
    elif masjed == StudentMasjedChoice.SALAM:
        sheet.cell(2, 2, 'السلام').alignment = alignment
    else:
        sheet.cell(2, 2, 'القزاز').alignment = alignment


    cell = sheet.cell(3, 1, 'تاريخ البداية')
    cell.alignment = alignment
    cell.font = header_font
    sheet.cell(3, 2, start_date).alignment = alignment

    cell = sheet.cell(4, 1, 'تاريخ النهاية')
    cell.alignment = alignment
    cell.font = header_font
    sheet.cell(4, 2, end_date).alignment = alignment

    cell = sheet.cell(5, 1, 'صفحات التسميع')
    cell.alignment = alignment
    cell.font = header_font
    sheet.cell(5, 2, data['total_memo']).alignment = alignment

    cell = sheet.cell(6, 1, 'صفحات السبر')
    cell.alignment = alignment
    cell.font = header_font
    sheet.cell(6, 2, data['total_test']).alignment = alignment

    cell = sheet.cell(7, 1, 'كلي الصفحات')
    cell.alignment = alignment
    cell.font = header_font
    sheet.cell(7, 2, data['total']).alignment = alignment

    details_cell_start = 9

    cell = sheet.cell(details_cell_start, 1, 'معرف الطالب')
    cell.alignment = alignment
    cell.font = header_font

    cell = sheet.cell(details_cell_start, 2, 'اسم الطالب')
    cell.alignment = alignment
    cell.font = header_font

    cell = sheet.cell(details_cell_start, 3, 'صفحات التسميع')
    cell.alignment = alignment
    cell.font = header_font

    cell = sheet.cell(details_cell_start, 4, 'صفحات السبر')
    cell.alignment = alignment
    cell.font = header_font

    cell = sheet.cell(details_cell_start, 5, 'كلي الصفحات')
    cell.alignment = alignment
    cell.font = header_font

    for row, student in enumerate(data['students'], details_cell_start + 1):
        student_id = student['student_id']
        student_name = student['student_name']
        sum_memo = student['sum_memo']
        sum_test = student['sum_test']
        sum_all = student['sum_all']

        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 24
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 12

        sheet.cell(row, 1, student_id).alignment = alignment
        sheet.cell(row, 2, student_name).alignment = alignment
        sheet.cell(row, 3, sum_memo).alignment = alignment
        sheet.cell(row, 4, sum_test).alignment = alignment
        sheet.cell(row, 5, sum_all).alignment = alignment

    buffer = BytesIO()

    workbook.save(buffer)

    buffer.seek(0)

    return buffer.getvalue()


def _get_masters_map():
    result_map = {}
    masters = get_user_model().objects.all()

    for master in masters:
        full_name = f"{master.first_name} {master.last_name}"
        result_map[master.pk] = full_name

    return result_map